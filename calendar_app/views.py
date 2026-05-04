import calendar
from datetime import date, datetime, timedelta, timezone as dt_timezone
from io import BytesIO
from urllib.parse import urlencode

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.colors import HexColor, white
from reportlab.pdfgen import canvas as rl_canvas

from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST

from .forms import AdminPanelFilterForm, CalendarFilterForm, EventForm, EventListFilterForm, EventQuickForm, unique_slug
from .models import Department, Event, Responsible


# Nombres de meses en español
MONTH_NAMES = [
    '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
    'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
]
DAY_NAMES = ['Dom', 'Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb']


def _spread_multiday_events(events_qs, year, month):
    """
    Construye events_by_day teniendo en cuenta eventos multi-día:
    un evento aparece en cada día del mes que abarca (inicio → fin).
    Cada entrada es un dict con la info del evento más metadatos de posición:
      is_start, is_end, is_continuation
    """
    month_start = date(year, month, 1)
    # Último día del mes
    if month == 12:
        month_end = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        month_end = date(year, month + 1, 1) - timedelta(days=1)

    events_by_day = {}
    for event in events_qs:
        # Usar hora local para evitar desfase por zona horaria UTC
        ev_start = timezone.localtime(event.start_datetime).date()
        ev_end = (timezone.localtime(event.end_datetime).date() if event.end_datetime else ev_start)

        # Rango visible dentro del mes
        visible_start = max(ev_start, month_start)
        visible_end = min(ev_end, month_end)

        cur = visible_start
        while cur <= visible_end:
            events_by_day.setdefault(cur.day, []).append({
                'event': event,
                'is_start': cur == ev_start,
                'is_end': cur == ev_end,
                'is_continuation': cur > ev_start,
                'is_multiday': ev_start != ev_end,
            })
            cur += timedelta(days=1)

    return events_by_day


def _is_mobile(request):
    ua = request.META.get('HTTP_USER_AGENT', '').lower()
    mobile_keywords = ('mobile', 'android', 'iphone', 'ipad', 'ipod', 'blackberry', 'windows phone', 'opera mini', 'opera mobi')
    return any(kw in ua for kw in mobile_keywords)


def calendar_view(request):
    if _is_mobile(request):
        return redirect(reverse('calendar_app:event_list') + ('?' + request.GET.urlencode() if request.GET else ''))

    today = date.today()

    filters = CalendarFilterForm(request.GET)
    filters.is_valid()  # limpia los campos válidos, ignora inválidos
    fd = filters.cleaned_data

    year = fd.get('year') or today.year
    month = fd.get('month') or today.month
    if month < 1:
        month, year = 12, year - 1
    elif month > 12:
        month, year = 1, year + 1

    dept_filters = fd.get('dept', [])
    scope_filters = fd.get('scope', [])
    type_filters = fd.get('type', [])

    events_qs = Event.objects.filter(
        # Incluye eventos que comienzan en el mes O que empiezan antes y terminan en el mes
        Q(start_datetime__year=year, start_datetime__month=month) |
        Q(start_datetime__date__lt=date(year, month, 1),
          end_datetime__date__gte=date(year, month, 1)),
        is_published=True,
    ).select_related('department', 'responsible')

    if dept_filters:
        events_qs = events_qs.filter(department__slug__in=dept_filters)
    if scope_filters:
        events_qs = events_qs.filter(scope__in=scope_filters)
    if type_filters:
        events_qs = events_qs.filter(event_type__in=type_filters)

    events_by_day = _spread_multiday_events(events_qs, year, month)

    cal = calendar.Calendar(firstweekday=6).monthdayscalendar(year, month)
    calendar_weeks = []
    for week in cal:
        days = []
        for day in week:
            if day == 0:
                days.append({'day': '', 'in_month': False, 'is_today': False, 'events': []})
            else:
                days.append({
                    'day': day,
                    'in_month': True,
                    'is_today': date(year, month, day) == today,
                    'events': events_by_day.get(day, []),
                })
        calendar_weeks.append(days)

    return render(request, 'calendar_app/calendar.html', {
        'year': year,
        'month': month,
        'month_name': MONTH_NAMES[month],
        'day_names': DAY_NAMES,
        'calendar_weeks': calendar_weeks,
        'departments': Department.objects.filter(is_active=True),
        'dept_filters': dept_filters,
        'scope_filters': scope_filters,
        'type_filters': type_filters,
        'today': today,
        'scope_options': [
            ('local', 'Local'),
            ('association', 'Asociación'),
            ('union', 'Unión'),
            ('general', 'General'),
        ],
        'event_type_options': [
            ('service', 'Culto'),
            ('seminar', 'Seminario'),
            ('meeting', 'Reunión'),
            ('social', 'Social'),
            ('youth', 'Jóvenes'),
            ('children', 'Infantil'),
            ('outreach', 'Evangelismo'),
            ('other', 'Otro'),
        ],
    })


def _build_event_list_qs(fd):
    """Aplica los filtros del EventListFilterForm a un queryset de eventos publicados."""
    qs = Event.objects.filter(is_published=True).select_related('department').order_by('start_datetime')
    if fd.get('q'):
        qs = qs.filter(
            Q(title__icontains=fd['q']) |
            Q(description__icontains=fd['q']) |
            Q(location__icontains=fd['q'])
        )
    if fd.get('dept'):
        qs = qs.filter(department__slug=fd['dept'])
    if fd.get('year'):
        qs = qs.filter(start_datetime__year=fd['year'])
    if fd.get('month'):
        qs = qs.filter(start_datetime__month=fd['month'])
    if fd.get('scope'):
        qs = qs.filter(scope=fd['scope'])
    return qs


def event_list(request):
    filters = EventListFilterForm(request.GET)
    filters.is_valid()
    fd = filters.cleaned_data

    qs = _build_event_list_qs(fd)

    # Rango de años disponibles (eventos existentes ± margen)
    today = date.today()
    year_range = list(range(today.year - 2, today.year + 3))

    # Querystring sin 'page' para paginación y exportación
    filter_params = {k: v for k, v in request.GET.items() if k != 'page'}
    filter_querystring = urlencode(filter_params)

    paginator = Paginator(qs, 20)
    events = paginator.get_page(request.GET.get('page'))

    return render(request, 'calendar_app/event_list.html', {
        'events': events,
        'filters': filters,
        'departments': Department.objects.filter(is_active=True),
        'year_range': year_range,
        'filter_querystring': filter_querystring,
        'scope_options': Event.ScopeChoices.choices,
        'month_names': MONTH_NAMES,
    })


def calendar_pdf_view(request):
    """Genera y descarga el calendario mensual como PDF usando ReportLab."""
    today = date.today()

    filters = CalendarFilterForm(request.GET)
    filters.is_valid()
    fd = filters.cleaned_data

    year = fd.get('year') or today.year
    month = fd.get('month') or today.month
    if month < 1:
        month, year = 12, year - 1
    elif month > 12:
        month, year = 1, year + 1

    dept_filters  = fd.get('dept', [])
    scope_filters = fd.get('scope', [])
    type_filters  = fd.get('type', [])

    events_qs = (
        Event.objects.filter(
            start_datetime__year=year,
            start_datetime__month=month,
            is_published=True,
        )
        .select_related('department')
        .order_by('start_datetime')
    )
    if dept_filters:
        events_qs = events_qs.filter(department__slug__in=dept_filters)
    if scope_filters:
        events_qs = events_qs.filter(scope__in=scope_filters)
    if type_filters:
        events_qs = events_qs.filter(event_type__in=type_filters)

    events_by_day = {}
    for ev in events_qs:
        d = timezone.localtime(ev.start_datetime).day
        events_by_day.setdefault(d, []).append(ev)

    cal_weeks = calendar.Calendar(firstweekday=6).monthdayscalendar(year, month)

    # ── PDF setup ──────────────────────────────────────────────────────
    buffer = BytesIO()
    page_w, page_h = landscape(A4)   # 841.89 x 595.28 pts
    c = rl_canvas.Canvas(buffer, pagesize=(page_w, page_h))
    c.setTitle(f'Calendario {MONTH_NAMES[month]} {year}')
    c.setAuthor('Iglesia Adventista Osorno Central')

    MX       = 18   # horizontal margin
    MT       = 16   # top margin
    MB       = 10   # bottom margin
    HEADER_H = 42
    DAYROW_H = 18

    usable_w  = page_w - 2 * MX
    col_w     = usable_w / 7
    num_rows  = len(cal_weeks)
    grid_h    = page_h - MT - MB - HEADER_H - DAYROW_H
    row_h     = grid_h / num_rows

    # ── Colors ─────────────────────────────────────────────────────────
    C_PRIMARY    = HexColor('#002045')
    C_TODAY_BG   = HexColor('#dde6f5')
    C_OFF_BG     = HexColor('#f4f5f7')
    C_BORDER     = HexColor('#d8dae4')
    C_DAYROW_BG  = HexColor('#eef0f5')
    C_TEXT_DARK  = HexColor('#1b1b1e')
    C_TEXT_MID   = HexColor('#74777f')
    C_TEXT_LIGHT = HexColor('#a8aab4')
    C_PILL_BG    = HexColor('#eff1f7')

    # ── Header ─────────────────────────────────────────────────────────
    base_y = page_h - MT

    c.setFont('Helvetica-Bold', 9)
    c.setFillColor(C_PRIMARY)
    c.drawString(MX, base_y - 11, 'CALENDARIO IGLESIA ADVENTISTA OSORNO CENTRAL')

    c.setFont('Helvetica-Bold', 20)
    c.setFillColor(C_PRIMARY)
    c.drawString(MX, base_y - 32, f'{MONTH_NAMES[month]} {year}')

    c.setFont('Helvetica', 6.5)
    c.setFillColor(C_TEXT_LIGHT)
    c.drawRightString(page_w - MX, base_y - 10, f'Generado el {today.strftime("%d/%m/%Y")}')

    sep_y = base_y - HEADER_H
    c.setStrokeColor(C_BORDER)
    c.setLineWidth(0.5)
    c.line(MX, sep_y, page_w - MX, sep_y)

    # ── Day names row ──────────────────────────────────────────────────
    dayrow_y = sep_y - DAYROW_H
    c.setFillColor(C_DAYROW_BG)
    c.rect(MX, dayrow_y, usable_w, DAYROW_H, fill=1, stroke=0)

    for i, dn in enumerate(['DOM', 'LUN', 'MAR', 'MIE', 'JUE', 'VIE', 'SAB']):
        cx = MX + i * col_w + col_w / 2
        c.setFillColor(C_PRIMARY if i in (0, 6) else C_TEXT_MID)
        c.setFont('Helvetica-Bold', 7)
        c.drawCentredString(cx, dayrow_y + 5, dn)

    # ── Calendar cells ─────────────────────────────────────────────────
    PILL_H   = 11
    PILL_GAP = 1.5
    FONT_EV  = 7

    for ri, week in enumerate(cal_weeks):
        for ci, day in enumerate(week):
            cx = MX + ci * col_w
            cy = dayrow_y - (ri + 1) * row_h
            in_month = day != 0
            is_today = in_month and date(year, month, day) == today
            is_wend  = ci in (0, 6)

            if not in_month:
                bg = C_OFF_BG
            elif is_today:
                bg = C_TODAY_BG
            else:
                bg = white

            c.setFillColor(bg)
            c.setStrokeColor(C_BORDER)
            c.setLineWidth(0.4)
            c.rect(cx, cy, col_w, row_h, fill=1, stroke=1)

            if not in_month:
                continue

            # Day number circle
            NUM_R  = 7.5
            num_cx = cx + col_w - NUM_R - 3
            num_cy = cy + row_h - NUM_R - 3

            if is_today:
                c.setFillColor(C_PRIMARY)
                c.circle(num_cx, num_cy, NUM_R, fill=1, stroke=0)
                c.setFillColor(white)
            else:
                c.setFillColor(C_PRIMARY if is_wend else C_TEXT_MID)

            c.setFont('Helvetica-Bold', 7.5)
            c.drawCentredString(num_cx, num_cy - 3, str(day))

            # Events
            evs       = events_by_day.get(day, [])
            pill_x    = cx + 3
            pill_w    = col_w - 6
            avail_h   = row_h - NUM_R * 2 - 9
            max_pills = max(1, int(avail_h / (PILL_H + PILL_GAP)))
            shown     = evs[:max_pills]
            overflow  = len(evs) - len(shown)
            pill_top  = cy + row_h - NUM_R * 2 - 8

            for pi, ev in enumerate(shown):
                py = pill_top - pi * (PILL_H + PILL_GAP) - PILL_H

                color_hex = ev.department.color if ev.department else '#94a3b8'
                try:
                    ev_color = HexColor(color_hex)
                except Exception:
                    ev_color = HexColor('#94a3b8')

                # Pill background
                c.setFillColor(C_PILL_BG)
                c.rect(pill_x, py, pill_w, PILL_H, fill=1, stroke=0)

                # Colored left stripe
                c.setFillColor(ev_color)
                c.rect(pill_x, py, 3, PILL_H, fill=1, stroke=0)

                # Event text
                local_dt = timezone.localtime(ev.start_datetime)
                prefix   = '' if ev.is_all_day else local_dt.strftime('%H:%M') + ' '
                text     = prefix + ev.title

                c.setFont('Helvetica', FONT_EV)
                max_w = pill_w - 7
                total_w = c.stringWidth(text, 'Helvetica', FONT_EV)
                if total_w > max_w and len(text) > 3:
                    cutoff = max(1, int(len(text) * max_w / total_w) - 1)
                    text = text[:cutoff].rstrip() + '...'
                    # corrección de un paso por si el estimado se pasó levemente
                    while len(text) > 4 and c.stringWidth(text, 'Helvetica', FONT_EV) > max_w:
                        text = text[:-4] + '...'

                c.setFillColor(C_TEXT_DARK)
                c.drawString(pill_x + 5, py + 2, text)

            if overflow > 0:
                ov_y = pill_top - len(shown) * (PILL_H + PILL_GAP) - 7
                c.setFont('Helvetica', 5)
                c.setFillColor(C_TEXT_LIGHT)
                c.drawString(pill_x + 2, ov_y, f'+{overflow} mas')

    # ── Outer border ───────────────────────────────────────────────────
    outer_h = DAYROW_H + num_rows * row_h
    c.setStrokeColor(HexColor('#b0b4c4'))
    c.setLineWidth(0.8)
    c.rect(MX, dayrow_y - num_rows * row_h, usable_w, outer_h, fill=0, stroke=1)

    c.save()

    month_ascii = (
        MONTH_NAMES[month].lower()
        .replace('é', 'e').replace('á', 'a').replace('ú', 'u')
        .replace('ó', 'o').replace('í', 'i')
    )
    filename = f'calendario-{month_ascii}-{year}.pdf'
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def event_list_export(request):
    filters = EventListFilterForm(request.GET)
    filters.is_valid()
    fd = filters.cleaned_data

    qs = _build_event_list_qs(fd)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Actividades'

    # ── Estilo encabezado ──────────────────────────────────────────────
    header_fill = PatternFill('solid', fgColor='002045')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = [
        ('Título', 40),
        ('Departamento', 22),
        ('Fecha inicio', 15),
        ('Hora inicio', 13),
        ('Fecha término', 15),
        ('Hora término', 13),
        ('Todo el día', 13),
        ('Lugar', 30),
        ('Alcance', 15),
        ('Tipo', 18),
        ('Público', 18),
        ('Requiere inscripción', 20),
        ('URL inscripción', 35),
        ('Publicado', 13),
    ]

    for col_idx, (title, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'

    # ── Filas de datos ─────────────────────────────────────────────────
    alt_fill = PatternFill('solid', fgColor='EFF4FF')
    data_align = Alignment(vertical='center', wrap_text=False)

    for row_idx, event in enumerate(qs, start=2):
        row_data = [
            event.title,
            event.department.name if event.department else '',
            event.start_datetime.strftime('%d/%m/%Y'),
            '' if event.is_all_day else event.start_datetime.strftime('%H:%M'),
            event.end_datetime.strftime('%d/%m/%Y') if event.end_datetime else '',
            '' if (event.is_all_day or not event.end_datetime) else event.end_datetime.strftime('%H:%M'),
            'Sí' if event.is_all_day else 'No',
            event.location,
            event.get_scope_display(),
            event.get_event_type_display(),
            event.get_audience_display(),
            'Sí' if event.requires_registration else 'No',
            event.registration_url,
            'Sí' if event.is_published else 'No',
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_align
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # ── Enviar respuesta ───────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="actividades.xlsx"'
    return response


def event_detail(request, slug):
    event = get_object_or_404(Event, slug=slug, is_published=True)

    # URL absoluta del evento (para compartir)
    event_url = request.build_absolute_uri(
        reverse('calendar_app:event_detail', args=[event.slug])
    )

    # URL de Google Calendar
    start_utc = event.start_datetime.astimezone(dt_timezone.utc)
    end_dt = event.end_datetime or (event.start_datetime + timedelta(hours=1))
    end_utc = end_dt.astimezone(dt_timezone.utc)
    fmt = '%Y%m%dT%H%M%SZ'
    gcal_params = urlencode({
        'action': 'TEMPLATE',
        'text': event.title,
        'dates': f"{start_utc.strftime(fmt)}/{end_utc.strftime(fmt)}",
        'details': event.description or '',
        'location': event.location or '',
    })
    google_cal_url = f"https://calendar.google.com/calendar/render?{gcal_params}"

    return render(request, 'calendar_app/event_detail.html', {
        'event': event,
        'event_url': event_url,
        'google_cal_url': google_cal_url,
        'departments': Department.objects.filter(is_active=True),
        'responsibles': Responsible.objects.filter(
            department__is_active=True
        ).select_related('department').order_by('full_name'),
    })


def tv_view(request):
    today = date.today()
    year = today.year
    month = today.month

    today_events = Event.objects.filter(
        start_datetime__date=today,
        is_published=True,
    ).select_related('department').order_by('start_datetime')

    upcoming_events = Event.objects.filter(
        start_datetime__date__gt=today,
        is_published=True,
    ).select_related('department').order_by('start_datetime')[:6]

    events_qs = Event.objects.filter(
        Q(start_datetime__year=year, start_datetime__month=month) |
        Q(start_datetime__date__lt=date(year, month, 1),
          end_datetime__date__gte=date(year, month, 1)),
        is_published=True,
    ).select_related('department')

    events_by_day = _spread_multiday_events(events_qs, year, month)

    # Domingo primero (firstweekday=6 = domingo en python-calendar)
    cal = calendar.Calendar(firstweekday=6).monthdayscalendar(year, month)
    calendar_weeks = []
    for week in cal:
        days = []
        for day in week:
            if day == 0:
                days.append({'day': '', 'in_month': False, 'is_today': False, 'events': []})
            else:
                days.append({
                    'day': day,
                    'in_month': True,
                    'is_today': date(year, month, day) == today,
                    'events': events_by_day.get(day, []),
                })
        calendar_weeks.append(days)

    DAY_NAMES_FULL_SUN = ['Domingo', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']
    MONTH_NAMES_FULL = [
        '', 'enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
        'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre',
    ]
    today_label = f"Sábado, {today.day} de {MONTH_NAMES_FULL[today.month]} de {today.year}" \
        if today.weekday() == 5 else \
        f"{DAY_NAMES_FULL_SUN[(today.weekday() + 1) % 7]}, {today.day} de {MONTH_NAMES_FULL[today.month]} de {today.year}"

    return render(request, 'calendar_app/tv.html', {
        'today': today,
        'today_label': today_label,
        'today_events': today_events,
        'upcoming_events': upcoming_events,
        'year': year,
        'month': month,
        'month_name': MONTH_NAMES[month],
        'day_names': DAY_NAMES_FULL_SUN,
        'calendar_weeks': calendar_weeks,
        'num_weeks': len(calendar_weeks),
    })


@staff_member_required
def admin_panel_view(request):
    today = date.today()

    # Gestión rápida: crear evento borrador
    if request.method == 'POST':
        form = EventQuickForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            event = Event.objects.create(
                title=cd['title'],
                slug=unique_slug(cd['title']),
                start_datetime=form.start_datetime,
                is_published=False,
                responsible=cd['responsible'],
            )
            messages.success(request, f'Actividad "{event.title}" creada. Completa los detalles.')
            return redirect('admin:calendar_app_event_change', event.pk)
        else:
            messages.error(request, 'Verifica los datos del formulario.')

    # Filtros de navegación del calendario
    panel_filters = AdminPanelFilterForm(request.GET)
    panel_filters.is_valid()
    fd = panel_filters.cleaned_data

    year = fd.get('year') or today.year
    month = fd.get('month') or today.month
    if month < 1:
        month, year = 12, year - 1
    elif month > 12:
        month, year = 1, year + 1

    dept_filter = fd.get('dept', '')
    scope_filter = fd.get('scope', '')

    events_qs = Event.objects.filter(
        Q(start_datetime__year=year, start_datetime__month=month) |
        Q(start_datetime__date__lt=date(year, month, 1),
          end_datetime__date__gte=date(year, month, 1)),
    ).select_related('department', 'responsible')

    if dept_filter:
        events_qs = events_qs.filter(department__slug=dept_filter)
    if scope_filter:
        events_qs = events_qs.filter(scope=scope_filter)

    events_by_day = _spread_multiday_events(events_qs, year, month)

    # Domingo primero
    sun_cal = calendar.Calendar(firstweekday=6)
    calendar_weeks = []
    for week in sun_cal.monthdayscalendar(year, month):
        days = []
        for day in week:
            if day == 0:
                days.append({'day': '', 'in_month': False, 'is_today': False, 'events': []})
            else:
                days.append({
                    'day': day,
                    'in_month': True,
                    'is_today': date(year, month, day) == today,
                    'events': events_by_day.get(day, []),
                })
        calendar_weeks.append(days)

    upcoming_events = Event.objects.filter(
        start_datetime__date__gte=today,
    ).select_related('department').order_by('start_datetime')[:8]

    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1

    return render(request, 'calendar_app/admin_panel.html', {
        'year': year,
        'month': month,
        'month_name': MONTH_NAMES[month],
        'day_names': ['Dom', 'Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb'],
        'calendar_weeks': calendar_weeks,
        'departments': Department.objects.filter(is_active=True),
        'responsibles': Responsible.objects.filter(
            department__is_active=True
        ).select_related('department').order_by('full_name'),
        'dept_filter': dept_filter,
        'scope_filter': scope_filter,
        'upcoming_events': upcoming_events,
        'prev_year': prev_year,
        'prev_month': prev_month,
        'next_year': next_year,
        'next_month': next_month,
        'today': today,
        'scope_choices': [
            ('', 'Todos los Alcances'),
            ('local', 'Local (Osorno Central)'),
            ('association', 'Asociación Sur Austral'),
            ('union', 'Unión'),
            ('general', 'General'),
        ],
    })


@require_POST
@staff_member_required
def event_create_view(request):
    """Crear un nuevo evento desde el modal del panel de secretaría."""
    form = EventForm(request.POST)
    if not form.is_valid():
        first_error = next(iter(form.errors.values()))[0]
        return JsonResponse({'ok': False, 'error': first_error}, status=400)

    cd = form.cleaned_data
    event = Event.objects.create(
        title=cd['title'],
        slug=unique_slug(cd['title']),
        description=cd['description'],
        department=cd['department'],
        responsible=cd['responsible'],
        scope=cd['scope'],
        event_type=cd['event_type'],
        audience=cd['audience'],
        location=cd['location'],
        bible_verse=cd['bible_verse'],
        is_all_day=cd['is_all_day'],
        start_datetime=form.start_datetime,
        end_datetime=form.end_datetime,
        is_published=cd['is_published'],
        is_featured=cd['is_featured'],
    )
    return JsonResponse({
        'ok': True,
        'title': event.title,
        'detail_url': reverse('calendar_app:event_detail', args=[event.slug]),
    })


def event_get_view(request, pk):
    """Devuelve datos JSON de un evento para rellenar el modal de edición."""
    event = get_object_or_404(Event, pk=pk)
    local_start = timezone.localtime(event.start_datetime)
    local_end = timezone.localtime(event.end_datetime) if event.end_datetime else None
    return JsonResponse({
        'ok': True,
        'id': event.pk,
        'title': event.title,
        'description': event.description or '',
        'department': event.department_id or '',
        'responsible': event.responsible_id or '',
        'location': event.location or '',
        'is_all_day': event.is_all_day,
        'start_date': local_start.strftime('%Y-%m-%d'),
        'start_time': local_start.strftime('%H:%M'),
        'end_date': local_end.strftime('%Y-%m-%d') if local_end else '',
        'end_time': local_end.strftime('%H:%M') if local_end else '',
        'event_type': event.event_type,
        'scope': event.scope,
        'audience': event.audience,
        'bible_verse': event.bible_verse or '',
        'is_published': event.is_published,
        'is_featured': event.is_featured,
    })


@require_POST
@staff_member_required
def event_update_view(request, pk):
    """Actualizar un evento existente desde el modal de edición."""
    event = get_object_or_404(Event, pk=pk)
    form = EventForm(request.POST)
    if not form.is_valid():
        first_error = next(iter(form.errors.values()))[0]
        return JsonResponse({'ok': False, 'error': first_error}, status=400)

    cd = form.cleaned_data
    if event.title != cd['title']:
        event.slug = unique_slug(cd['title'], exclude_pk=pk)

    event.title = cd['title']
    event.description = cd['description']
    event.department = cd['department']
    event.responsible = cd['responsible']
    event.scope = cd['scope']
    event.event_type = cd['event_type']
    event.audience = cd['audience']
    event.location = cd['location']
    event.bible_verse = cd['bible_verse']
    event.is_all_day = cd['is_all_day']
    event.start_datetime = form.start_datetime
    event.end_datetime = form.end_datetime
    event.is_published = cd['is_published']
    event.is_featured = cd['is_featured']
    event.save()
    return JsonResponse({
        'ok': True,
        'title': event.title,
        'detail_url': reverse('calendar_app:event_detail', args=[event.slug]),
    })


@require_POST
@staff_member_required
def event_delete_view(request, pk):
    """Eliminar un evento."""
    event = get_object_or_404(Event, pk=pk)
    title = event.title
    event.delete()
    return JsonResponse({'ok': True, 'title': title})


def event_ical_view(request, slug):
    """Descarga el evento en formato iCalendar (.ics) para Apple Calendar, Outlook, etc."""
    event = get_object_or_404(Event, slug=slug, is_published=True)

    fmt = '%Y%m%dT%H%M%SZ'
    now_str = datetime.now(tz=dt_timezone.utc).strftime(fmt)
    start_utc = event.start_datetime.astimezone(dt_timezone.utc)
    end_dt = event.end_datetime or (event.start_datetime + timedelta(hours=1))
    end_utc = end_dt.astimezone(dt_timezone.utc)

    event_url = request.build_absolute_uri(
        reverse('calendar_app:event_detail', args=[event.slug])
    )

    def ical_text(value):
        """Escapa caracteres especiales para iCal."""
        return (value or '').replace('\\', '\\\\').replace('\n', '\\n').replace(',', '\\,').replace(';', '\\;')

    ical = (
        "BEGIN:VCALENDAR\r\n"
        "VERSION:2.0\r\n"
        "PRODID:-//Iglesia Adventista Osorno Central//ES\r\n"
        "CALSCALE:GREGORIAN\r\n"
        "METHOD:PUBLISH\r\n"
        "BEGIN:VEVENT\r\n"
        f"UID:{event.slug}@osorno-central\r\n"
        f"DTSTAMP:{now_str}\r\n"
        f"DTSTART:{start_utc.strftime(fmt)}\r\n"
        f"DTEND:{end_utc.strftime(fmt)}\r\n"
        f"SUMMARY:{ical_text(event.title)}\r\n"
        f"DESCRIPTION:{ical_text(event.description)}\r\n"
        f"LOCATION:{ical_text(event.location)}\r\n"
        f"URL:{event_url}\r\n"
        "END:VEVENT\r\n"
        "END:VCALENDAR\r\n"
    )

    response = HttpResponse(ical, content_type='text/calendar; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{event.slug}.ics"'
    return response
